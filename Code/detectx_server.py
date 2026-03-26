from flask import Flask, request, jsonify, send_from_directory, after_this_request
import cv2, numpy as np, base64, os, json, pandas as pd
from ultralytics import YOLO
import mediapipe as mp
import easyocr
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

EXCEL_FILE  = "detectx_logs.xlsx"
SNAP_DIR    = "violation_snapshots"
YOLO_MODEL  = "yolov8n.pt"
FINE_TRIPLE = 1000
FINE_HELMET = 1000

os.makedirs(SNAP_DIR, exist_ok=True)

app = Flask(__name__, static_folder=".", static_url_path="")

@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response

@app.route("/analyze", methods=["OPTIONS"])
@app.route("/ping",    methods=["OPTIONS"])
def options_handler():
    return "", 204

print("[INFO] Loading YOLOv8...")
yolo = YOLO(YOLO_MODEL)

print("[INFO] Loading EasyOCR...")
reader = easyocr.Reader(['en'], gpu=False)

print("[INFO] Loading MediaPipe...")
mp_face       = mp.solutions.face_detection
face_detector = mp_face.FaceDetection(min_detection_confidence=0.5)

print("[INFO] All models ready. Server starting...\n")


def count_riders(bike_box, people_boxes):
    bx1,by1,bx2,by2 = bike_box
    riders = [b for b in people_boxes if b[0]<bx2 and b[2]>bx1 and b[1]<by2 and b[3]>by1]
    return len(riders), riders

def check_helmet(frame, person_box):
    x1,y1,x2,y2 = person_box
    head_y2 = y1 + int((y2-y1)*0.4)
    crop = frame[y1:head_y2, x1:x2]
    if crop.size == 0: return True
    result = face_detector.process(cv2.cvtColor(crop, cv2.COLOR_BGR2RGB))
    return result.detections is None

def ocr_region(region):
    best_text, best_conf = "", 0.0
    gray    = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
    resized = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    blurred = cv2.GaussianBlur(resized, (3,3), 0)
    _, otsu = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    adaptive = cv2.adaptiveThreshold(blurred,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,11,2)
    for img in [otsu, adaptive, resized]:
        try:
            for (_,text,conf) in reader.readtext(img):
                clean = text.strip()
                if len(''.join(c for c in clean if c.isalnum()))>=4 and conf>best_conf:
                    best_conf, best_text = conf, clean
        except Exception: pass
    return best_text or None

def detect_plate(frame):
    h,w = frame.shape[:2]
    gray    = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray,(5,5),0)
    canny   = cv2.Canny(blurred,50,150)
    sx = cv2.Sobel(blurred,cv2.CV_8U,1,0,ksize=3)
    sy = cv2.Sobel(blurred,cv2.CV_8U,0,1,ksize=3)
    combined = cv2.dilate(cv2.bitwise_or(canny,cv2.bitwise_or(sx,sy)),
                          cv2.getStructuringElement(cv2.MORPH_RECT,(5,3)),iterations=1)
    contours,_ = cv2.findContours(combined,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    cands = sorted(
        [(x,y,cw,ch) for cnt in contours
         for x,y,cw,ch in [cv2.boundingRect(cnt)]
         if 1.5<(cw/ch if ch else 0)<6.0 and 1500<cw*ch<w*h*0.08 and cw>60 and ch>15],
        key=lambda c:c[2]*c[3], reverse=True)
    for x,y,cw,ch in cands[:10]:
        region = frame[max(0,y-5):min(h,y+ch+5), max(0,x-5):min(w,x+cw+5)]
        t = ocr_region(region)
        if t: return t
    t = ocr_region(frame[h//2:, :])
    return t if t else "Not Detected"

def frame_to_b64(frame):
    _, buf = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 88])
    return base64.b64encode(buf).decode()

def img_to_b64(path):
    try:
        with open(path,"rb") as f: return base64.b64encode(f.read()).decode()
    except: return ""

def draw_label(frame, text, pos, color=(0,255,255), scale=0.6, thick=2):
    x,y = pos
    (tw,th),_ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, scale, thick)
    cv2.rectangle(frame,(x,y-th-4),(x+tw+4,y+4),(0,0,0),-1)
    cv2.putText(frame,text,(x+2,y),cv2.FONT_HERSHEY_SIMPLEX,scale,color,thick)

def save_to_excel(data, snapshot_path):
    df_new = pd.DataFrame([data])
    if os.path.exists(EXCEL_FILE):
        try: df_new = pd.concat([pd.read_excel(EXCEL_FILE), df_new], ignore_index=True)
        except: pass
    target = EXCEL_FILE
    try: df_new.to_excel(target, index=False)
    except PermissionError:
        target = EXCEL_FILE.replace(".xlsx",f"_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        df_new.to_excel(target, index=False)
    try:
        wb = load_workbook(target); ws = wb.active; row = ws.max_row
        img_col = len(data)+1; col_l = get_column_letter(img_col)
        ws.column_dimensions[col_l].width = 26
        ws.row_dimensions[row].height = 80
        if row == 2: ws.cell(row=1,column=img_col,value="Snapshot")
        if os.path.exists(snapshot_path):
            thumb = os.path.splitext(snapshot_path)[0]+"_thumb.jpg"
            cv2.imwrite(thumb, cv2.resize(cv2.imread(snapshot_path),(180,90)))
            xl = XLImage(thumb); xl.width=180; xl.height=90
            ws.add_image(xl, f"{col_l}{row}")
        wb.save(target)
    except Exception as e:
        print(f"[WARN] Excel image embed: {e}")


def run_detection(frame):
    results_yolo  = yolo(frame, verbose=False)
    people_boxes, bike_boxes = [], []

    for r in results_yolo:
        for box in r.boxes:
            cls          = int(box.cls[0])
            x1,y1,x2,y2 = map(int, box.xyxy[0])
            conf         = float(box.conf[0])
            if cls == 0:
                people_boxes.append((x1,y1,x2,y2))
                cv2.rectangle(frame,(x1,y1),(x2,y2),(0,255,0),2)
                draw_label(frame,f"Person {conf:.0%}",(x1,max(y1-5,15)),(0,255,0))
            elif cls == 3:
                bike_boxes.append((x1,y1,x2,y2))
                cv2.rectangle(frame,(x1,y1),(x2,y2),(255,80,0),2)

    plate_number   = detect_plate(frame)
    violations_out = []

    for idx, bike in enumerate(bike_boxes):
        bx1,by1,bx2,by2 = bike
        riders_count, riders = count_riders(bike, people_boxes)
        violations, fine = [], 0
        helmet = check_helmet(frame, riders[0]) if riders else True

        if riders_count > 2: violations.append("Triple Riding"); fine += FINE_TRIPLE
        if not helmet:        violations.append("No Helmet");     fine += FINE_HELMET

        ly = max(by1-70, 70)
        draw_label(frame,f"Riders: {riders_count}",              (bx1,ly),    (0,255,255))
        draw_label(frame,f"Helmet: {'YES' if helmet else 'NO'}", (bx1,ly+22), (0,200,0) if helmet else (0,0,255))
        draw_label(frame,f"Plate: {plate_number}",               (bx1,ly+44), (255,255,0))

        snap_b64 = ""
        if violations:
            draw_label(frame," | ".join(violations)+f"  Fine: Rs.{fine}",
                       (bx1,by2+22),(0,0,255),scale=0.65,thick=2)

            fh,fw = frame.shape[:2]; pad=30
            crop = frame[max(0,by1-pad):min(fh,by2+pad), max(0,bx1-pad):min(fw,bx2+pad)]
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            snap_path = os.path.join(SNAP_DIR, f"violation_{ts}_{idx}.jpg")
            cv2.imwrite(snap_path, crop)
            snap_b64 = img_to_b64(snap_path)

            now = datetime.now()
            save_to_excel({
                "Date":now.strftime("%Y-%m-%d"),"Time":now.strftime("%H:%M:%S"),
                "Plate Number":plate_number,"Violation":", ".join(violations),
                "Rider Count":riders_count,"Helmet":"Yes" if helmet else "No",
                "Fine (Rs.)":fine,"Snapshot File":snap_path
            }, snap_path)

            violations_out.append({
                "bike_index":  idx,
                "plate":       plate_number,
                "plate_visible": plate_number != "Not Detected",
                "no_helmet":   not helmet,
                "overloaded":  riders_count > 2,
                "rider_count": riders_count,
                "fine":        fine,
                "violations":  violations,
                "description": ", ".join(violations),
                "snapshot_b64": snap_b64,
                "time": datetime.now().strftime("%H:%M:%S"),
                "date": datetime.now().strftime("%Y-%m-%d"),
            })

    return {
        "output_image_b64": frame_to_b64(frame),
        "violations":       violations_out,
        "total_fine":       sum(v["fine"] for v in violations_out),
        "plate":            plate_number,
        "scene_summary":    f"{len(violations_out)} violation(s) detected. Plate: {plate_number}.",
        "processed_at":     datetime.now().isoformat(),
    }


@app.route("/")
def index():
    return send_from_directory(".", "detectx_dashboard.html")

@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        file = request.files.get("image")
        if not file:
            return jsonify({"error": "No image uploaded"}), 400

        # Decode image
        img_bytes = np.frombuffer(file.read(), np.uint8)
        frame     = cv2.imdecode(img_bytes, cv2.IMREAD_COLOR)
        if frame is None:
            return jsonify({"error": "Could not decode image"}), 400

        # Resize if too large
        h,w = frame.shape[:2]
        if w > 1600:
            frame = cv2.resize(frame, (1600, int(h*1600/w)))

        result = run_detection(frame)
        return jsonify(result)

    except Exception as e:
        print(f"[ERROR] {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/ping")
def ping():
    return jsonify({"status": "ok"})

if __name__ == "__main__":
    print("=" * 52)
    print("  DetectX Server running!")
    print("  Open http://localhost:5000 in your browser")
    print("=" * 52)
    app.run(host="0.0.0.0", port=5000, debug=False)
